from django.shortcuts import render, redirect,HttpResponse
from app01.models import *
from app01.utils.pagination import Pagination

# Create your views here.
def depart_list(request):
    """ 部门列表 """
    # 去数据库获取数据
    queryset = Department.objects.all()
    page_obj = Pagination(request,queryset,page_size=2)
    set = {'queryset':page_obj.page_queryset,
           'page_string':page_obj.html()}
    return render(request,'depart_list.html',set)

def depart_add(request):
    """ 添加部门 """
    if request.method == 'GET':
        return render(request,'depart_add.html')
    title = request.POST.get('title')
    Department.objects.create(title=title)
    return redirect('/depart/list/')

def depart_delete(request):
    """ 删除 部门 """

    nid = request.GET.get('nid')
    Department.objects.filter(id = nid).delete()
    return redirect('/depart/list/')

def depart_edit(request,nid):
    """修改部门"""
    if request.method == 'GET':
        # 根据nid获取到部门数据
        row_object = Department.objects.filter(id=nid).first()
        return render(request,'depart_edit.html',{'row_object':row_object})
    title = request.POST.get('title')
    #  根据id找到数据库数据并更新
    Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')

def depart_mutil(request):
    ''' 文件上传 (excel)'''
    # 获取用户上传文件对象
    file_obj = request.FILES.get('exc')
    # 打开并读取内容 通过openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]
    # 循环获取每行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        # print(text)
        # 此时就可以创建数据库数据
        if not Department.objects.filter(title=text).exists():
            Department.objects.create(title=text)


    # cell = sheet.cell(1,1)
    # print(cell.value)
    return redirect('/depart/list/')
